# -*- coding: utf-8 -*-
"""Build the CoreTrace GUI ATP workbook from the Epitech template."""
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

TEMPLATE = r"C:\Users\shookapic\Downloads\ATP_template.xlsx"
OUT = r"c:\Users\shookapic\Documents\CoretraceProjetcs\coretrace-gui\atp\CoreTrace-GUI_ATP.xlsx"

REL = "https://github.com/CoreTrace/coretrace-gui/releases/download/v5.1.0"

ACCESS = [
    ("Release under test", "https://github.com/CoreTrace/coretrace-gui/releases/tag/v5.1.0 (v5.1.0)"),
    ("Linux AppImage", REL + "/CtraceGUI-5.1.0.AppImage"),
    ("Windows installer", REL + "/CtraceGUI-Setup-5.1.0.exe"),
    ("macOS, Apple Silicon (DMG)", REL + "/CtraceGUI-5.1.0-arm64.dmg"),
    ("macOS, Apple Silicon (ZIP)", REL + "/CtraceGUI-5.1.0-arm64-mac.zip"),
    ("macOS, Intel", "Not published in v5.1.0 (the x64 build job is being fixed). Intel testers use the Docker environment below."),
    ("Docker test environment", "appendices/docker/ in the appendices ZIP: run ./build.sh then ./run.sh, then open http://localhost:6080/"),
    ("Test data set", "appendices/datasets/ in the appendices ZIP (expected results: datasets/MANIFEST.md)"),
    ("Tester instructions", "appendices/README.md in the appendices ZIP"),
    ("Source code", "https://github.com/CoreTrace/coretrace-gui"),
    ("API documentation", "https://coretrace.github.io/coretrace-gui/"),
    ("Anomaly reports", "https://github.com/CoreTrace/coretrace-gui/issues"),
]

DATASETS = [
    ("DS01 - stack buffer overflow", "DS01_stack_overflow.c",
     "1 WARNING - StackBufferOverflow on 'buf' (size 16), line 12, index may go up to 32"),
    ("DS02 - clean file", "DS02_clean.c",
     "0 diagnostic - the analysis succeeds and the empty-results state is shown"),
    ("DS03 - out-of-bounds index", "DS03_index_oob.c",
     "2 WARNINGs - StackBufferOverflow on 'tab' (size 10) line 9, UninitializedLocalRead (CWE-457) line 10"),
    ("DS04 - unreachable path", "DS04_infeasible_path.c",
     "0 diagnostic - the overflow sits on a path the guard makes unreachable (no false positive)"),
    ("DS05 - infinite recursion", "DS05_recursion.c",
     "1 INFO + 1 ERROR - recursive function detected, then unconditional self recursion (no base case), line 9"),
    ("DS06 - C++ input", "DS06_cpp_array_oob.cpp",
     "2 WARNINGs - StackBufferOverflow on 'tab' (size 8) line 10, UninitializedLocalRead line 12"),
    ("DS07 - unsupported input", "DS07_invalid.txt",
     "The analysis fails and the panel reports the error (Unsupported input file type)"),
    ("DS08 - large source file", "DS08_large_file.c",
     "About 2 MB: the editor shows the first 1 MB and a 'Load next 1 MB' button"),
    ("AI assistant credentials", "not provided",
     "The tester supplies their own API key (OpenAI, Anthropic, Groq, Deepseek or Perplexity) or a local .gguf model file"),
]

CRIT, SEC, UI = "Critical feature", "Secondary function", "User interface"
DONE, WIP = "Done", "In progress"

S = [
("F1", "Launch on Linux through Docker", CRIT,
 "Bring the reference test environment up from the published binary, with no build step.",
 "Docker installed and running, 3 GB of free disk space, an internet connection.",
 "1. Unzip the appendices archive\n2. cd appendices/docker\n3. Run ./build.sh (downloads the 506 MB AppImage, about 5 min)\n4. Run ./run.sh\n5. Open http://localhost:6080/ in a browser (it connects on its own)",
 "The build ends with 'Image built: coretrace-gui-atp'. The browser shows the CoreTrace GUI window: custom title bar, activity bar on the left, Explorer sidebar, and a welcome screen in the middle. The status bar bottom-right reads 'CtraceGUI v5.1.0'.",
 DONE),
("F2", "Launch on Linux from the AppImage", CRIT,
 "Check the published Linux artifact starts on a plain desktop, outside Docker.",
 "An x86-64 Linux desktop session, FUSE 2 available (otherwise use the documented --appimage-extract fallback).",
 "1. Download CtraceGUI-5.1.0.AppImage from the release page\n2. chmod +x CtraceGUI-5.1.0.AppImage\n3. ./CtraceGUI-5.1.0.AppImage",
 "The application window opens within 30 seconds and shows the same interface as F1, with 'CtraceGUI v5.1.0' in the status bar. No error dialog is displayed.",
 DONE),
("F3", "Install and launch on macOS (Apple Silicon)", CRIT,
 "Install the unsigned macOS build and get past Gatekeeper without help.",
 "A Mac with Apple Silicon (M1 or later) running macOS 12 or newer.",
 "1. Download CtraceGUI-5.1.0-arm64.dmg from the release page\n2. Open the .dmg and drag CtraceGUI to Applications\n3. Double-click the app: macOS refuses to open it\n4. Right-click (or Control-click) CtraceGUI in Applications and choose Open, then click Open in the dialog\n5. If step 4 reported 'is damaged and can't be opened' instead, run in Terminal: xattr -cr /Applications/CtraceGUI.app  then launch the app normally\n6. Quit the app and launch it again by double-clicking",
 "Step 3 is refused by Gatekeeper because the build is unsigned; that refusal is expected, not a failure. After step 4 (or 5) the application window opens and shows the same interface as F1, with 'CtraceGUI v5.1.0' in the status bar. Step 6 opens the app directly, with no further warning.",
 DONE),
("F3b", "Launch on macOS, Intel", CRIT,
 "Give Intel Macs a way to run the plan while the x64 build is missing.",
 "An Intel Mac with Docker Desktop installed and at least 6 GB of RAM allocated to it.",
 "1. cd appendices/docker\n2. Run ./build.sh\n3. Run ./run.sh\n4. Open http://localhost:6080/ (it connects on its own)",
 "The CoreTrace GUI window is displayed in the browser and is usable, and every other scenario of this plan can be executed in it. v5.1.0 publishes no Intel artifact, so this Docker path is the supported one on Intel until the x64 build job is fixed.",
 WIP),
("F4", "Backend version indicator", SEC,
 "Show which CoreTrace analysis backend the application is aware of.",
 "The application is running (F1, F2 or F3) with an internet connection.",
 "1. Look at the bottom-right of the status bar\n2. Hover the 'CoreTrace latest: ...' label",
 "The label first reads 'CoreTrace latest: checking...' then settles on a release tag or on an explicit status. It never stays on 'checking...' for more than 30 seconds. The tooltip reads 'Latest CoreTrace backend release: <tag>'.",
 DONE),
("F5", "Open a folder as a workspace", CRIT,
 "Load the test data set into the file explorer.",
 "The application is running; the datasets folder is reachable (in Docker it is mounted at /home/tester/workspace).",
 "1. Menu File > Open Folder...\n2. Select the datasets folder (/home/tester/workspace in Docker)\n3. Confirm",
 "The sidebar switches from 'No folder opened' to the workspace root named after the folder, and lists the 9 data set entries (DS01 to DS08 plus MANIFEST.md) sorted with folders first. The Explorer icon in the activity bar stays highlighted.",
 DONE),
("F6", "Expand a subfolder", SEC,
 "Check directories load their contents on demand.",
 "A workspace containing at least one subfolder is open (create one with F10 if needed).",
 "1. Click the arrow of a folder in the tree\n2. Click it again to collapse it",
 "The folder expands and lists its children indented below it; the arrow flips. Collapsing hides them again. The rest of the tree is unchanged.",
 DONE),
("F7", "Open a file in the editor", CRIT,
 "Display the content of a source file.",
 "The datasets workspace is open (F5).",
 "1. Click DS01_stack_overflow.c in the file tree",
 "A tab named DS01_stack_overflow.c appears above the editor, the welcome screen is replaced by the source code with line numbers, and the status bar shows 'Ln 1, Col 1', 'UTF-8' and a C/C++ file type.",
 DONE),
("F8", "Work with several open files", CRIT,
 "Edit more than one file at a time.",
 "The datasets workspace is open.",
 "1. Open DS01_stack_overflow.c\n2. Open DS02_clean.c\n3. Open DS03_index_oob.c\n4. Click the first tab\n5. Click the x on the second tab",
 "Three tabs are listed; clicking a tab shows that file's content and marks it active. Closing a tab removes it and leaves the other two open, with one of them active.",
 DONE),
("F9", "Create a file", CRIT,
 "Create a new source file inside the workspace.",
 "The datasets workspace is open.",
 "1. Right-click an empty area of the file tree\n2. Choose 'New file..'\n3. Type test_atp.c and confirm",
 "test_atp.c appears in the tree, opens in an empty editor tab, and exists on disk (visible from the integrated terminal with ls).",
 DONE),
("F10", "Create a folder", SEC,
 "Create a directory inside the workspace.",
 "The datasets workspace is open.",
 "1. Right-click an empty area of the file tree\n2. Choose 'New folder..'\n3. Type atp_subdir and confirm",
 "atp_subdir appears in the tree as a collapsible folder, listed before the files.",
 DONE),
("F11", "Rename a file", SEC,
 "Rename an item from the explorer.",
 "test_atp.c exists in the workspace (F9).",
 "1. Right-click test_atp.c\n2. Choose 'Rename'\n3. Type test_atp_renamed.c and confirm",
 "The tree entry is renamed, the old name is gone, and the open tab (if any) follows the new name. No duplicate entry is left behind.",
 DONE),
("F12", "Delete a file", SEC,
 "Remove an item from the explorer.",
 "test_atp_renamed.c exists in the workspace (F11).",
 "1. Right-click test_atp_renamed.c\n2. Choose 'Delete'\n3. Confirm the deletion",
 "The file disappears from the tree and from disk. If it was open, its tab is closed or flagged as missing. No other file is affected.",
 DONE),
("F13", "Save a modified file", CRIT,
 "Persist an edit to disk.",
 "DS02_clean.c is open in the editor.",
 "1. Type a comment line at the top of DS02_clean.c\n2. Menu File > Save\n3. Close the tab and reopen the file",
 "The tab loses its unsaved-changes marker on save, a confirmation notification appears, and reopening the file shows the added comment.",
 DONE),
("F14", "Save As", SEC,
 "Write the current buffer to a different path.",
 "DS02_clean.c is open with at least one modification.",
 "1. Menu File > Save As...\n2. Choose the workspace folder and the name copy_of_DS02.c\n3. Confirm",
 "copy_of_DS02.c is created with the current editor content, the active tab now points to the new file, and the original DS02_clean.c is left untouched on disk.",
 DONE),
("F15", "Auto Save", SEC,
 "Save modifications without an explicit command.",
 "A file is open in the editor.",
 "1. Menu File > Auto Save\n2. Check the 'Auto Save: ON' indicator in the status bar\n3. Type a few characters and wait 5 seconds without saving\n4. Reopen the file",
 "The status bar shows the Auto Save indicator as ON, and the typed characters are present after reopening the file. Selecting Auto Save again turns the indicator off.",
 DONE),
("F16", "Detect an external change", SEC,
 "Keep the explorer in sync with the file system.",
 "The datasets workspace is open and the integrated terminal is available (F39).",
 "1. Open the integrated terminal\n2. Run: touch external_change.c\n3. Look at the file tree without clicking anything",
 "external_change.c appears in the tree within a few seconds without a manual refresh.",
 DONE),
("F17", "Refresh the file tree", UI,
 "Force a reload of the workspace listing.",
 "The datasets workspace is open.",
 "1. Click the refresh icon at the top of the Explorer panel",
 "The tree is rebuilt: the listing is identical apart from any change made on disk, and no open tab is lost.",
 DONE),
("F18", "Partial loading of a large file", SEC,
 "Keep the editor responsive on a file larger than 1 MB.",
 "The datasets workspace is open.",
 "1. Open DS08_large_file.c (about 2 MB)\n2. Read the banner shown under the editor\n3. Click 'Load next 1 MB'",
 "The file opens without freezing the interface, a banner reads 'Showing first 1 MB' with a 'Load next 1 MB' button, and clicking it appends the following megabyte to the editor content.",
 DONE),
("F19", "File deleted while open", SEC,
 "Fail safely when the file backing a tab disappears.",
 "copy_of_DS02.c is open in a tab (F14).",
 "1. In the integrated terminal, run: rm copy_of_DS02.c\n2. Look at the editor\n3. Open 'Ctrace Tools' and click 'Run Analysis'",
 "A banner reads 'File not found on disk.' with a 'Close tab' button. The analysis is refused with 'File not found - analysis blocked' instead of running on a missing file. The application does not crash.",
 DONE),
("F20", "Close the workspace", SEC,
 "Return to the empty state.",
 "A workspace is open.",
 "1. Menu File > Close Folder (or right-click the workspace root, then Close Folder)",
 "The tree is emptied and the sidebar shows the 'no folder opened' state again. The application stays responsive.",
 DONE),
("F21", "C/C++ syntax highlighting", UI,
 "Make source code readable in the editor.",
 "The datasets workspace is open.",
 "1. Open DS06_cpp_array_oob.cpp\n2. Open DS07_invalid.txt in a second tab",
 "In the .cpp file, keywords (int, for, return), the #include directive, strings and comments are coloured differently from plain identifiers, and the status bar shows a C++ file type. The .txt file is shown unhighlighted as plain text.",
 DONE),
("F22", "Find in the current file", SEC,
 "Locate an occurrence inside the open file.",
 "DS03_index_oob.c is open.",
 "1. Menu Edit > Find\n2. Type tab\n3. Press Enter twice",
 "The find widget opens, matches are highlighted, the occurrence counter shows the number of hits, and Enter moves the selection from one occurrence to the next.",
 DONE),
("F23", "Go to line", SEC,
 "Jump to a precise location, as reported by the analysis.",
 "DS03_index_oob.c is open.",
 "1. Menu Edit > Go to Line...\n2. Type 9 and confirm",
 "The editor scrolls to line 9 (tab[i] = i;), the cursor is placed on it, and the status bar reads 'Ln 9, Col 1'.",
 DONE),
("F25", "Toggle word wrap", UI,
 "Read long lines without scrolling horizontally.",
 "DS08_large_file.c is open.",
 "1. Menu Edit > Toggle Word Wrap\n2. Trigger it a second time",
 "The first activation wraps long lines inside the viewport and removes the horizontal scrollbar; the second restores the original single-line display.",
 DONE),
("F26", "Cursor position indicator", UI,
 "Report the caret position in the status bar.",
 "DS01_stack_overflow.c is open.",
 "1. Click in the middle of line 5\n2. Press the down arrow twice",
 "The status bar shows 'Ln 5, Col <n>' matching the clicked position, and updates to 'Ln 7, ...' after the two key presses.",
 DONE),
("F27", "Search across the workspace", CRIT,
 "Find a symbol in every file of the project.",
 "The datasets workspace is open.",
 "1. Click the Search icon in the activity bar\n2. Type StackBufferOverflow in the search field and confirm\n3. Click one of the results",
 "The panel lists the matching files with the matching lines (MANIFEST.md at least). Clicking a result opens that file in a tab, scrolled to the matching line.",
 DONE),
("F27b", "Analysis toolchain check", CRIT,
 "Prove the analyser can actually compile before trusting any analysis result.",
 "The datasets workspace is open. On Docker (F1) clang 20 is already installed; on a native AppImage (F2) it must be installed first, as appendices/README.md describes.",
 "1. Open the integrated terminal (View > Terminal)\n2. Run: clang --version\n3. Open DS05_recursion.c in the active tab\n4. Open 'Ctrace Tools' and click 'Run Analysis' with the default arguments",
 "Step 2 reports version 20.x. Step 4 reports 1 error and 1 info. If step 4 reports zero diagnostics, the toolchain is wrong and every later analysis scenario is meaningless: the analyser returns 'no diagnostics' rather than an error when it cannot compile. Stop and fix clang before running F28 to F38, and record the clang version in the Feedback column.",
 DONE),
("F28", "Analyse a file with a known overflow", CRIT,
 "Verify the core feature: reporting a real stack buffer overflow.",
 "The datasets workspace is open, DS01_stack_overflow.c open in the active tab.",
 "1. Menu 'Ctrace Tools'\n2. Leave the arguments field at its default value: --invoke=ctrace_stack_analyzer --sarif-format\n3. Click 'Run Analysis'\n4. Wait for the spinner to disappear (up to 60 s)",
 "The panel shows exactly 1 diagnostic, of severity WARNING, rule StackBufferOverflow, located at line 12 of DS01_stack_overflow.c. Its message names the variable 'buf' (size 16) and states that the index may go up to 32 while the last valid index is 15, on a write access.",
 DONE),
("F29", "Analyse a clean file", CRIT,
 "Verify the tool stays silent on correct code.",
 "The datasets workspace is open.",
 "1. Open DS02_clean.c in the active tab\n2. Open 'Ctrace Tools' and click 'Run Analysis' with the default arguments",
 "The analysis completes without error and the panel reports zero diagnostic (empty-results state). No warning or error entry is listed.",
 DONE),
("F30", "Several diagnostics of different rules", CRIT,
 "Check grouping and counting when a file has more than one defect.",
 "The datasets workspace is open.",
 "1. Open DS03_index_oob.c in the active tab\n2. Run the analysis with the default arguments",
 "The panel reports 2 WARNINGs, grouped by rule: StackBufferOverflow on 'tab' (size 10) at line 9, and UninitializedLocalRead (CWE-457, confidence 0.90) at line 10. The summary at the top of the panel counts 2 warnings.",
 DONE),
("F31", "No false positive on an unreachable path", CRIT,
 "Check that path refinement discards a defect that cannot be reached.",
 "The datasets workspace is open.",
 "1. Open DS04_infeasible_path.c and read the code: the write buf[i] is guarded by n < 8 then n > 16\n2. Run the analysis with the default arguments",
 "No StackBufferOverflow is reported, even though the write is syntactically out of bounds: the guard makes it unreachable. Zero diagnostic is listed.",
 DONE),
("F32", "Report an error-level defect", CRIT,
 "Check the highest severity is produced and displayed as such.",
 "The datasets workspace is open.",
 "1. Open DS05_recursion.c in the active tab\n2. Run the analysis with the default arguments",
 "Two entries are listed for line 9: one INFO ('recursive or mutually recursive function detected') and one ERROR ('unconditional self recursion detected (no base case)', with 'this will eventually overflow the stack at runtime'). The ERROR entry is visually distinct from the INFO one, and the summary counts 1 error.",
 DONE),
("F33", "Analyse a C++ file", CRIT,
 "Check .cpp inputs are supported, not only C.",
 "The datasets workspace is open.",
 "1. Open DS06_cpp_array_oob.cpp in the active tab\n2. Run the analysis with the default arguments",
 "2 WARNINGs are reported: StackBufferOverflow on 'tab' (size 8) at line 10, and UninitializedLocalRead at line 12. The results are rendered the same way as for a C file.",
 DONE),
("F34", "Unsupported input file", CRIT,
 "Fail explicitly rather than silently on a file the analyser cannot read.",
 "The datasets workspace is open.",
 "1. Open DS07_invalid.txt in the active tab\n2. Run the analysis with the default arguments",
 "The analysis terminates and the panel reports the failure, quoting the analyser's message ('Unsupported input file type' / 'Failed to analyze'). The application does not hang and stays usable afterwards.",
 DONE),
("F35", "Analysis without an open file", SEC,
 "Guard the analysis against an empty editor.",
 "The application is running with no file open (close every tab).",
 "1. Menu 'Ctrace Tools'\n2. Click 'Run Analysis'",
 "The panel shows 'No active file to analyze' with the hint 'Please open a file first', and a notification reads 'Open a file to analyze with CTrace'. Nothing is executed.",
 DONE),
("F36", "Custom analysis arguments", SEC,
 "Check the arguments field is passed through to the analyser.",
 "DS01_stack_overflow.c is open in the active tab.",
 "1. Open 'Ctrace Tools'\n2. Replace the arguments with: --invoke=notatool --sarif-format\n3. Click 'Run Analysis'\n4. Restore the default arguments and run again",
 "The first run fails and the panel surfaces the analyser's message listing the allowed tools (flawfinder, ikos, cppcheck, tscancode, ctrace_stack_analyzer). The second run produces the F28 result again, proving the field is really forwarded.",
 DONE),
("F37", "Jump from a diagnostic to the code", CRIT,
 "Navigate from a result to the offending line.",
 "The F30 analysis has been run on DS03_index_oob.c and its results are displayed.",
 "1. Click the StackBufferOverflow entry in the results panel",
 "The editor scrolls to line 9 of DS03_index_oob.c and highlights it; the status bar reads 'Ln 9, ...'. If the file is not the active tab, it is made active first.",
 DONE),
("F38", "Clear the results", UI,
 "Reset the analysis panel.",
 "Analysis results are displayed.",
 "1. Click the 'Clear' button next to 'Run Analysis'",
 "The results list is emptied and the panel returns to its placeholder ('Run CTrace to analyze your code'). The open files and tabs are unaffected.",
 DONE),
("F39", "Integrated terminal", SEC,
 "Run a shell command without leaving the application.",
 "The application is running.",
 "1. Click the Terminal icon in the activity bar (or menu View > Terminal)\n2. Type: ls -la\n3. Press Enter",
 "A terminal panel opens at the bottom, the command runs in the workspace folder, and its output is displayed. The prompt comes back ready for the next command.",
 DONE),
("F40", "Several terminals", SEC,
 "Manage more than one shell session.",
 "The terminal panel is open (F39).",
 "1. Click 'New Terminal'\n2. Switch between the two terminal tabs\n3. Click 'Kill Terminal' on the second one\n4. Close the panel with its x button",
 "A second terminal tab is created and both keep their own history; switching shows the right one. Killing removes only the selected terminal. Closing the panel hides it without affecting the editor.",
 DONE),
("F41", "Configure an AI provider", SEC,
 "Connect the assistant to a language model.",
 "The tester has an API key for one of the supported providers (OpenAI, Anthropic, Groq, Deepseek, Perplexity).",
 "1. Menu 'Assistant'\n2. Open the assistant settings\n3. Select a provider and paste the API key\n4. Pick a model from the list\n5. Click Save",
 "The model list is fetched after the key is entered (the status shows 'Fetching models...' then a populated list). After Save, the assistant panel is ready to accept a message and the key is not shown in clear text anywhere in the interface.",
 DONE),
("F42", "Ask about the current file", SEC,
 "Use the open source file as the assistant's context.",
 "A provider is configured (F41) and DS01_stack_overflow.c is open.",
 "1. Open the Assistant panel\n2. Click the 'Current file' context button\n3. Type: what is wrong with this code?\n4. Press Enter",
 "The context button becomes active and names the current file, a 'Thinking...' placeholder appears, then an answer is returned that refers to the buffer written past its bounds. Code blocks in the answer carry Copy and Replace buttons.",
 DONE),
("F43", "Ask about the analysis results", SEC,
 "Feed the diagnostics to the assistant.",
 "A provider is configured and the F28 analysis has been run.",
 "1. Open the Assistant panel\n2. Click the 'Analysis' context button\n3. Ask: how do I fix the reported warning?",
 "The context button reflects that results are attached (it reads 'No results' only when none exist), and the answer refers to the StackBufferOverflow reported on 'buf'.",
 DONE),
("F44", "Conversation history", SEC,
 "Find a previous exchange again.",
 "At least one exchange has taken place (F42).",
 "1. Click the '+' button to start a new conversation\n2. Send a message in it\n3. Click the history button\n4. Reopen the first conversation\n5. Delete the second one from the list",
 "The new conversation starts empty; the history lists both conversations; reopening one restores its full message list; deleting removes it from the list and it does not come back after a restart.",
 DONE),
("F45", "Stop a generation", SEC,
 "Interrupt a long answer.",
 "A provider is configured.",
 "1. Ask the assistant a question with a long answer (for example: explain static analysis in detail)\n2. While the answer is being produced, click the stop button",
 "Generation stops within a couple of seconds, the partial answer stays visible, the stop button is replaced by the send button, and a new message can be sent immediately.",
 DONE),
("F46", "Backend settings", SEC,
 "Point the application at a specific analyser binary.",
 "The application is running.",
 "1. Menu File > Backend Settings...\n2. Read the detected binary path and its status\n3. Use the browse button to select a binary, then cancel the dialog\n4. Close the settings",
 "The dialog reports the bundled ctrace binary and whether it was found. The file picker opens and cancelling leaves the current setting unchanged. After closing, an analysis (F28) still produces its expected result.",
 DONE),
("F47", "Update settings", SEC,
 "Choose the update channel and check for a new version.",
 "The application is running with an internet connection.",
 "1. Menu File > Update Settings...\n2. Switch the channel to 'beta'\n3. Trigger 'Check now'\n4. Switch back to 'main' and close",
 "The selected channel is saved and shown as selected when the dialog is reopened. 'Check now' reports a result (update available, up to date, or an explicit error) within 30 seconds instead of staying silent. No update is installed without an explicit confirmation.",
 WIP),
("F48", "Session restore", CRIT,
 "Do not lose work when the application is restarted.",
 "The datasets workspace is open with DS01_stack_overflow.c and DS03_index_oob.c in two tabs.",
 "1. Note the open tabs and the active one\n2. Close the application\n3. Start it again (./run.sh, or the AppImage)",
 "On restart the same workspace is loaded, the same two tabs are reopened in the same order, and the tab that was active is active again. A tab whose file was deleted meanwhile is flagged as missing rather than opened empty.",
 DONE),
("F49", "Interface layout controls", UI,
 "Adapt the layout to the task at hand.",
 "The application is running.",
 "1. Menu View > Toggle Sidebar, then again\n2. Switch between the Explorer and Search icons in the activity bar\n3. Menu View > Visual Effects\n4. Drag the border between the sidebar and the editor",
 "The sidebar hides and comes back with its content intact; the activity bar highlights the selected panel; the Visual Effects entry flips between On and Off and the change is visible; dragging resizes the sidebar and the editor follows without overlap.",
 DONE),
("F50", "Window controls", UI,
 "Check the custom title bar behaves like a native one.",
 "The application is running.",
 "1. Click the maximize button in the title bar\n2. Click it again to restore\n3. Click minimize, then bring the window back\n4. Click the close button",
 "Maximize fills the screen and restore returns to the previous size; minimize hides the window and it can be restored; close ends the application. In the Docker environment, minimize and maximize act on the noVNC canvas, and the container stops after close.",
 DONE),
("F51", "Welcome screen actions", UI,
 "Start working from a cold launch, with no workspace open.",
 "The application is running with no folder open and every tab closed.",
 "1. Read the welcome screen in the middle of the window\n2. Click its 'Open Folder' action and select the datasets folder\n3. Close the workspace (File > Close Folder) to bring the welcome screen back\n4. Click its 'New File' action",
 "The welcome screen shows 'Welcome to CTraceGUI' with three actions: New File, Open File and Open Folder. Open Folder loads the workspace exactly as F5 does. New File opens an empty untitled tab and replaces the welcome screen.",
 DONE),
("F52", "Switch tabs from the keyboard", SEC,
 "Move between open files without the mouse.",
 "DS01_stack_overflow.c, DS02_clean.c and DS03_index_oob.c are open in three tabs.",
 "1. Press Ctrl+Tab twice\n2. Press Ctrl+PageDown\n3. Press Ctrl+PageUp",
 "Each Ctrl+Tab activates the next tab in order, wrapping round after the last one. Ctrl+PageDown moves to the next tab and Ctrl+PageUp back to the previous one. The editor content and the status bar follow the active tab every time. These stay on Ctrl on macOS too, because Cmd+Tab is the system application switcher.",
 DONE),
("F53", "macOS keyboard shortcuts use Cmd", SEC,
 "Check the shortcuts follow the platform convention on a Mac.",
 "The application is running natively on macOS (F3), with the datasets workspace open.",
 "1. Open the File menu and read the shortcut hints\n2. Press Cmd+O, then cancel the dialog\n3. Open a file, type a character, press Cmd+S\n4. Press Cmd+F, then Escape\n5. Press Cmd+B twice\n6. Press Option+Z twice",
 "Menu hints read as macOS symbols (⌘N, ⌘O, ⌘⇧O, ⌘S, ⌘⇧S, ⌘W) rather than 'Ctrl+...'. Cmd+O opens the file dialog, Cmd+S saves and clears the unsaved marker, Cmd+F opens the find widget, Cmd+B hides and restores the sidebar, and Option+Z toggles word wrap. This scenario is not applicable on Linux.",
 DONE),
("F54", "Analysis is refused on macOS without a native binary", CRIT,
 "Fail with a usable message rather than an opaque one when the analyser cannot run.",
 "The application is running natively on macOS (F3), with DS01_stack_overflow.c open in the active tab, and no ctrace-darwin binary installed.",
 "1. Open 'Ctrace Tools'\n2. Click 'Run Analysis' with the default arguments",
 "The analysis is refused with a message stating that the bundled ctrace binary is a Linux executable and cannot run on macOS, and naming both remedies: select a macOS build from File > Backend Settings, or place one next to the bundled binary as 'ctrace-darwin-arm64'. The application does not hang or crash, and every other feature stays usable. This scenario is not applicable on Linux, where the bundled binary runs natively.",
 DONE),
("F55", "Git decorations in the file tree", SEC,
 "Show which files in the workspace have uncommitted changes.",
 "The datasets folder has been turned into a git repository: run 'git init && git add -A && git commit -m init' in the integrated terminal first.",
 "1. Refresh the file tree\n2. In the integrated terminal, run: echo '// edit' >> DS02_clean.c\n3. Refresh the file tree again",
 "After step 1 no file is decorated. After step 3, DS02_clean.c is marked as modified in the tree and its parent folders carry the roll-up marker. A workspace that is not a git repository shows no decoration and raises no error.",
 DONE),
("F56", "Performance overlay", UI,
 "Inspect rendering cost while using the application.",
 "The application is running with a workspace open.",
 "1. Press Ctrl+Alt+P (Cmd+Option+P on macOS)\n2. Read the values displayed\n3. Click 'Lite effects'\n4. Click 'Hide'",
 "An overlay titled 'Performance' appears with live FPS, Frame, Long tasks, DOM nodes, JS heap, GPU, Effects and DPR readings; FPS and Frame change as the window is scrolled. 'Lite effects' toggles the reduced-effects mode and the Effects value follows it. 'Hide' dismisses the overlay, and the shortcut brings it back.",
 DONE),
]

wb = openpyxl.load_workbook(TEMPLATE)

# ---------------------------------------------------------------- Info & Access
info = wb["Info & Access"]
for row in info.iter_rows(min_row=1, max_row=info.max_row + 5, max_col=4):
    for c in row:
        c.value = None
info.tables.clear()

info["A1"] = "Access to testing platforms"
info["A1"].font = Font(bold=True, size=12)
info["A2"], info["B2"] = "Acess point", "URL"
r = 3
for name, url in ACCESS:
    info.cell(row=r, column=1, value=name)
    info.cell(row=r, column=2, value=url)
    r += 1
access_end = r - 1

ds_label = access_end + 2
info.cell(row=ds_label, column=1, value="Data sets").font = Font(bold=True, size=12)
hdr = ds_label + 1
info.cell(row=hdr, column=1, value="Data set")
info.cell(row=hdr, column=2, value="File (appendices/datasets/)")
info.cell(row=hdr, column=3, value="Expected result")
r = hdr + 1
for a, b, c in DATASETS:
    info.cell(row=r, column=1, value=a)
    info.cell(row=r, column=2, value=b)
    info.cell(row=r, column=3, value=c)
    r += 1
ds_end = r - 1

note = ds_end + 2
info.cell(row=note, column=1, value="Notes").font = Font(bold=True, size=12)
NOTES = [
    "Version under test: CoreTrace GUI v5.1.0. Every expected result in the test plan was produced with the analysis binary bundled in that release.",
    "Linux is tested through the Docker image in appendices/docker (reference environment) or directly from the AppImage.",
    "The analyser requires clang 20: it compiles each file to LLVM IR before analysing it. Without clang 20 it reports 0 diagnostics for every input instead of failing, so run scenario F27b before F28 to F38. The Docker image already contains clang 20; a native AppImage does not.",
    "macOS, Apple Silicon: install the unsigned .dmg natively (scenario F3). The build is not signed, so Gatekeeper blocks the first launch on purpose — open it once with right-click > Open, or run 'xattr -cr /Applications/CtraceGUI.app'.",
    "macOS, Intel: v5.1.0 publishes no x64 artifact, so Intel Macs run the Docker image instead (scenario F3b).",
    "Analysis does not run on macOS: the bundled ctrace is a Linux binary. The application says so explicitly (scenario F54), and every other feature works. Run the analysis scenarios F28 to F38 on Linux.",
    "Shortcuts follow the platform: Cmd on macOS, Ctrl elsewhere (scenario F53). Tab navigation stays on Ctrl everywhere, because Cmd+Tab is the macOS application switcher.",
    "A scenario is validated when every expected result is obtained with no blocking or critical anomaly. Report anything else in the Feedback column of the test plan.",
]
for i, line in enumerate(NOTES):
    info.cell(row=note + 1 + i, column=1, value=line)

t1 = Table(displayName="AccessPoints", ref="A2:B%d" % access_end)
t1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
info.add_table(t1)
t2 = Table(displayName="DataSets", ref="A%d:C%d" % (hdr, ds_end))
t2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
info.add_table(t2)

info.column_dimensions["A"].width = 34
info.column_dimensions["B"].width = 62
info.column_dimensions["C"].width = 80
for row in info.iter_rows(min_row=1, max_row=info.max_row, max_col=3):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)

# ------------------------------------------------------------------- Test Plan
tp = wb["Test Plan"]
tbl = tp.tables["Tableau7"]
for row in tp.iter_rows(min_row=2, max_row=tp.max_row, max_col=9):
    for c in row:
        c.value = None
for i, sc in enumerate(S):
    for j, val in enumerate(sc):
        tp.cell(row=2 + i, column=1 + j, value=val)
last = 1 + len(S)
tbl.ref = "A1:I%d" % last
tbl.autoFilter.ref = tbl.ref

for row in tp.iter_rows(min_row=1, max_row=last, max_col=9):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)
for col, w in zip("ABCDEFGHI", [10, 34, 19, 44, 40, 52, 78, 18, 34]):
    tp.column_dimensions[col].width = w
tp.freeze_panes = "C2"

wb.save(OUT)
print("written:", OUT)
print("scenarios:", len(S), "| access rows:", len(ACCESS), "| datasets:", len(DATASETS))
